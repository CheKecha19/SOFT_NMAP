import re
import os
import sys
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from config import COLORS, DEFAULT_OUTPUT_DIR, OUTPUT_SUFFIX

def parse_nmap_txt(file_path):
    hosts = []
    scan_info = {
        'start_time': None,
        'command': None,
        'total_ips': 0,
        'hosts_up': 0
    }
    
    current_host = None
    host_pattern = re.compile(r'Nmap scan report for (?:([\w\-. ]+)\s)?\(?([\d.]+)\)?')
    ports_section = False
    
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            
            # Парсинг общей информации
            if not scan_info['start_time'] and line.startswith('Starting Nmap'):
                scan_info['start_time'] = line.split(' at ')[-1]
            
            elif 'Nmap done:' in line:
                match = re.search(r'Nmap done: (\d+) IP addresses \((\d+) hosts? up\)', line)
                if match:
                    scan_info['total_ips'] = int(match.group(1))
                    scan_info['hosts_up'] = int(match.group(2))
            
            # Парсинг хостов
            host_match = host_pattern.match(line)
            if host_match:
                if current_host:
                    hosts.append(current_host)
                
                hostname, ip = host_match.groups()
                current_host = {
                    'ip': ip,
                    'hostname': hostname.strip() if hostname else None,
                    'ports': {}
                }
                ports_section = False
                continue
            
            if current_host is None:
                continue
            
            # Начало секции с портами
            if line.startswith('PORT') and 'STATE' in line and 'SERVICE' in line:
                ports_section = True
                continue
            
            # Конец секции с портами
            if ports_section and (not line or line.startswith('Nmap scan')):
                ports_section = False
            
            # Парсинг портов
            if ports_section and line:
                if line.startswith('Not shown:') or line.startswith('All ') or 'filtered' in line:
                    continue
                
                parts = re.split(r'\s+', line, 2)
                if len(parts) < 2:
                    continue
                    
                port_service = parts[0].strip()
                state = parts[1].strip()
                service = parts[2].strip().split()[0] if len(parts) > 2 else 'unknown'
                
                if '/' not in port_service:
                    continue
                    
                port, protocol = port_service.split('/', 1)
                port_key = f"{port}/{service}"
                
                current_host['ports'][port_key] = state.lower()

    # Добавляем последний хост
    if current_host:
        hosts.append(current_host)
    
    return hosts, scan_info

def create_excel_report(hosts, scan_info, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "Scan Results"
    
    # Стили
    header_font = Font(bold=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    color_fills = {k: PatternFill(start_color=v, end_color=v, fill_type='solid') for k, v in COLORS.items()}
    
    # Тонкие границы для всех ячеек
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Заголовок
    ws.append(["Общая информация"])
    ws.append(["Время запуска сканирования:", scan_info.get('start_time', 'N/A')])
    ws.append(["Хост-источник:", "N/A"])
    ws.append(["Выполненная команда:", scan_info.get('command', 'N/A')])
    ws.append(["Обработано хостов:", f"{scan_info.get('hosts_up', 0)} (из {scan_info.get('total_ips', 0)})"])
    ws.append([])
    
    # Заголовки хостов
    headers = ["Хосты:"] + [host['ip'] for host in hosts]
    ws.append(headers)
    
    # Hostnames
    hostnames = ["hostname:"] + [host.get('hostname', '') for host in hosts]
    ws.append(hostnames)
    
    # Заголовок портов
    ws.append(["Порты:"])
    
    # Собираем все уникальные порты
    all_ports = sorted(set(
        port 
        for host in hosts 
        for port in host['ports'].keys()
    ))
    
    # Добавляем порты
    for port in all_ports:
        row = [port]
        for host in hosts:
            state = host['ports'].get(port, '')
            row.append(state if state else '')
        ws.append(row)
    
    # Применяем форматирование
    # 1. Объединяем только нужные ячейки
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))  # "Общая информация"
    ws.merge_cells(start_row=9, start_column=1, end_row=9, end_column=len(headers))  # "Порты:"
    
    # 2. Форматирование заголовков
    for row in ws.iter_rows(min_row=1, max_row=9 + len(all_ports), max_col=len(headers)):
        for cell in row:
            cell.font = header_font if cell.row <= 9 else Font()
            cell.border = thin_border
            if cell.row == 1 or cell.row == 9:
                cell.alignment = alignment_center
    
    # 3. Добавляем цвета для состояний портов
    for row_idx in range(10, 10 + len(all_ports)):
        port_cell = ws.cell(row=row_idx, column=1)
        port_cell.font = header_font
        
        for col_idx in range(2, len(headers) + 1):
            state_cell = ws.cell(row=row_idx, column=col_idx)
            state = state_cell.value.lower() if state_cell.value else ''
            
            if 'open' in state:
                state_cell.fill = color_fills['open']
            elif 'closed' in state:
                state_cell.fill = color_fills['closed']
            elif 'filtered' in state:
                state_cell.fill = color_fills['filtered']
            else:
                state_cell.fill = color_fills['default']
    
    # 4. Настраиваем ширину колонок
    for col_idx in range(1, len(headers) + 1):
        column_letter = get_column_letter(col_idx)
        max_length = 0
        
        for row in ws.iter_rows(min_row=1, min_col=col_idx, max_col=col_idx):
            cell = row[0]
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        adjusted_width = max(15, min(50, max_length + 2))
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 5. Центрирование текста в заголовках хостов
    for col in range(1, len(headers) + 1):
        for row in [7, 8]:  # Строки с заголовками хостов и hostname
            ws.cell(row=row, column=col).alignment = alignment_center
    
    # Сохраняем файл
    wb.save(output_file)
    print(f"Отчёт сохранён как: {output_file}")

def main():
    if len(sys.argv) < 2:
        print("Использование: py zenmap.py <путь_к_файлу.txt> [папка_выхода]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    output_dir = sys.argv[2] if len(sys.argv) > 2 else DEFAULT_OUTPUT_DIR
    
    if not os.path.exists(input_file):
        print(f"Ошибка: файл '{input_file}' не найден")
        sys.exit(1)
    
    base_name = os.path.splitext(os.path.basename(input_file))[0]
    output_file_name = f"{base_name}{OUTPUT_SUFFIX}"
    
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, output_file_name)
    else:
        output_path = os.path.join(os.path.dirname(__file__), output_file_name)
    
    try:
        hosts, scan_info = parse_nmap_txt(input_file)
        
        if not hosts:
            print("Предупреждение: не найдено данных для обработки")
            wb = Workbook()
            wb.save(output_path)
            print(f"Создан пустой отчет: {output_path}")
        else:
            create_excel_report(hosts, scan_info, output_path)
            
    except Exception as e:
        print(f"Критическая ошибка при обработке файла: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(2)

if __name__ == "__main__":
    main()
